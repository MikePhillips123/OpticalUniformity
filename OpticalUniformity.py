from dearpygui.dearpygui import *
import dearpygui.dearpygui as dpg
from dearpygui._dearpygui import configure_item as _configure_item
import tkinter as tk
from tkinter import filedialog
import os
import cv2
import numpy as np
import pandas as pd
import shutil
import threading
from openpyxl.styles import Font, Border, Side
from datetime import datetime
import sys

###### CONFIGURATION ######

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(__file__)
RESOURCE_DIR = os.path.join(BASE_DIR, "resources")
OUTPUT_DIR = os.path.join(BASE_DIR, "Output")
EXCEL_PATH = os.path.join(BASE_DIR, "Output", "Data")
GRID_SIZE = 5
CROP_SIZE = 1500
theme_counter = 0
analysis_running = False
last_selected_path = None

###### FILE HANDLING ######

def save_settings(sender, app_data, user_data):
    settings = {
        # Thick
        "color_r_input_1": get_value("color_r_input_1"),
        "color_g_input_1": get_value("color_g_input_1"),
        "color_b_input_1": get_value("color_b_input_1"),

        # Even
        "color_r_input_2": get_value("color_r_input_2"),
        "color_g_input_2": get_value("color_g_input_2"),
        "color_b_input_2": get_value("color_b_input_2"),

        # Thin
        "color_r_input_3": get_value("color_r_input_3"),
        "color_g_input_3": get_value("color_g_input_3"),
        "color_b_input_3": get_value("color_b_input_3"),

        # Legend layout settings
        "swatch_size": get_value("swatch_size"),
        "entry_padding": get_value("entry_padding"),
        "swatch_text_padding": get_value("swatch_text_padding"),

        # Legend font and background
        "legend_bg_r": get_value("legend_bg_r"),
        "legend_bg_g": get_value("legend_bg_g"),
        "legend_bg_b": get_value("legend_bg_b"),
        "legend_font_r": get_value("legend_font_r"),
        "legend_font_g": get_value("legend_font_g"),
        "legend_font_b": get_value("legend_font_b"),

        # Font settings
        "legend_font_size": get_value("legend_font_size"),
        "legend_font_thickness": get_value("legend_font_thickness"),
        "crop_position_selector": get_value("crop_position_selector"),

    }

    with open("image_output_settings.txt", "w") as f:
        for key, value in settings.items():
            f.write(f"{key}={value}\n")

    log("Settings saved.")

def load_settings():
    try:
        with open("image_output_settings.txt", "r") as f:
            for line in f:
                key, value = line.strip().split("=")
                if does_item_exist(key):
                    try:
                        # Try setting as float
                        set_value(key, float(value))
                    except ValueError:
                        # Try setting as bool
                        if value.lower() in ['true', 'false']:
                            set_value(key, value.lower() == 'true')
                        else:
                            # Fallback to string (e.g., for combo boxes like crop_position_selector)
                            set_value(key, value)

    except FileNotFoundError:
        log("Settings file not found. Using defaults.")

    update_color_preview(None, None, ("color_r_input_1", "color_g_input_1", "color_b_input_1", "preview_1"))
    update_color_preview(None, None, ("color_r_input_2", "color_g_input_2", "color_b_input_2", "preview_2"))
    update_color_preview(None, None, ("color_r_input_3", "color_g_input_3", "color_b_input_3", "preview_3"))

def save_log_to_file():
    try:
        log_lines = []
        children = get_item_children("log_window", 1)  # 1 = item type
        if children:
            for child in children:
                text = get_value(child)
                log_lines.append(text)

        if not log_lines:
            print("[INFO] No logs to save.")
            return

        # Create log directory if it doesn't exist
        log_dir = os.path.join(BASE_DIR, "Log")
        os.makedirs(log_dir, exist_ok=True)

        timestamp = datetime.now().strftime("Log [%d-%m-%y] [%H.%M].txt")
        log_path = os.path.join(log_dir, timestamp)

        with open(log_path, "w", encoding="utf-8") as f:
            for line in log_lines:
                f.write(line + "\n")

        log(f"[INFO] Log saved to {log_path}")
    except Exception as e:
        log(f"[ERROR] Failed to save log: {e}")

###### GUI HANDLING ######

def analyse_images_entry(sender, app_data, user_data):
    global analysis_running
    if not analysis_running:
        analysis_running = True
        disable_item("analyse_button_settings")
        disable_item("analyse_button_directories")
        threading.Thread(target=analyse_images, daemon=True).start()
    else:
        log("[INFO] Image analysis is already running.")


def open_color_picker(sender, app_data, user_data):
    set_value("active_rgb_inputs", str(user_data))  # store as string

    if does_item_exist("Color Picker Window"):
        delete_item("Color Picker Window")

    with window(label="Color Picker", tag="Color Picker Window", width=350, height=350, modal=True, no_resize=False):
        add_text("Select a color:")
        add_color_picker(tag="color_picker_value", default_value=[0.0, 0.0, 0.0, 1.0], no_alpha=True, display_rgb=True, width=250)
        add_spacer()
        add_button(label="OK", callback=confirm_color_selection)

def confirm_color_selection(sender, app_data, user_data):
    rgba = get_value("color_picker_value")
    if rgba is None:
        log("No color selected")
        return
    rgb = [int(c) for c in rgba[:3]]
    log(f"Selected RGB values: {rgb}")
    r_tag, g_tag, b_tag = eval(get_value("active_rgb_inputs"))  # read and convert

    set_value(r_tag, rgb[0])
    set_value(g_tag, rgb[1])
    set_value(b_tag, rgb[2])

    if r_tag == "color_r_input_1":
        preview_tag = "preview_1"
    elif r_tag == "color_r_input_2":
        preview_tag = "preview_2"
    elif r_tag == "color_r_input_3":
        preview_tag = "preview_3"
    elif r_tag == "legend_bg_r":
        preview_tag = "legend_bg_preview"
    elif r_tag == "legend_font_r":
        preview_tag = "legend_font_preview"
    else:
        preview_tag = None

    if preview_tag:
        update_color_preview(None, None, (r_tag, g_tag, b_tag, preview_tag))
        update_font_preview()
    if preview_tag in ("legend_font_preview", "legend_bg_preview"):
        update_font_preview()

    log(f"Raw RGBA from picker: {rgba}")

    delete_item("Color Picker Window")

def open_folder_dialog(sender, app_data, user_data):
    global last_selected_path  # ⬅ enable modifying the global

    root = tk.Tk()
    root.withdraw()

    # First time: use home dir
    if last_selected_path and os.path.isdir(last_selected_path):
        initial_dir = os.path.dirname(last_selected_path)  # Go up one level
    else:
        initial_dir = os.path.expanduser("~")

    path = filedialog.askdirectory(title="Select folder", initialdir=initial_dir)
    root.destroy()

    if path:
        last_selected_path = path  # store for next time
        update_output_preview(path)

        path_tag = user_data["path_tag"]
        folder_tag = user_data["folder_tag"]
        folder_name = os.path.basename(path)

        if does_item_exist(path_tag):
            set_value(path_tag, path)
        if does_item_exist(folder_tag):
            set_value(folder_tag, folder_name)

def folder_selected(sender, app_data, user_data):
    tags = get_item_user_data("folder_picker")  # Grab the data set earlier
    path_tag = tags["path_tag"]
    folder_tag = tags["folder_tag"]

    path = app_data['file_path_name']
    folder_name = path.rstrip("/\\").split("/")[-1]  # Extract folder name

    if does_item_exist(path_tag):
        set_value(path_tag, path)
    if does_item_exist(folder_tag):
        set_value(folder_tag, folder_name)

def update_color_preview(sender, app_data, user_data):
    global theme_counter
    r_tag, g_tag, b_tag, preview_tag = user_data
    r, g, b = get_value(r_tag), get_value(g_tag), get_value(b_tag)
    theme_counter += 1
    theme_tag = f"{preview_tag}_theme_{theme_counter}"
    with theme(tag=theme_tag):
        with theme_component(mvButton):
            add_theme_color(mvThemeCol_Button, (r, g, b, 255), category=mvThemeCat_Core)
            add_theme_color(mvThemeCol_ButtonHovered, (r, g, b, 255), category=mvThemeCat_Core)
            add_theme_color(mvThemeCol_ButtonActive, (r, g, b, 255), category=mvThemeCat_Core)
    bind_item_theme(preview_tag, theme_tag)

    if last_selected_path:
        update_output_preview(last_selected_path)

def update_font_preview():
    if last_selected_path:
        update_output_preview(last_selected_path)

def generate_legend_preview_image():
    font_scale = dpg.get_value("legend_font_size") / 32.0
    font_thickness = int(round(get_value("legend_font_thickness")))

    bg_color = (
        dpg.get_value("legend_bg_b"),  # B
        dpg.get_value("legend_bg_g"),  # G
        dpg.get_value("legend_bg_r")  # R
    )

    font_color = (
        dpg.get_value("legend_font_b"),
        dpg.get_value("legend_font_g"),
        dpg.get_value("legend_font_r")
    )

    entries = [
        ("Thin", (
            dpg.get_value("color_b_input_3"),
            dpg.get_value("color_g_input_3"),
            dpg.get_value("color_r_input_3")
        )),
        ("Even", (
            dpg.get_value("color_b_input_2"),
            dpg.get_value("color_g_input_2"),
            dpg.get_value("color_r_input_2")
        )),
        ("Thick", (
            dpg.get_value("color_b_input_1"),
            dpg.get_value("color_g_input_1"),
            dpg.get_value("color_r_input_1")
        )),
    ]

    # Step 1: upscale factor for drawing
    scale_factor = 2  # 2× resolution
    legend_width = 600 * scale_factor

    # Step 2: create a larger dummy image
    dummy_img = np.zeros((1, legend_width, 3), dtype=np.uint8)

    # Step 3: draw high-res version of legend
    hires_img = add_legend_below(
        dummy_img,
        entries=entries,
        pad=int(dpg.get_value("entry_padding") * scale_factor),
        swatch_size=int(dpg.get_value("swatch_size") * scale_factor),
        line_spacing=10 * scale_factor,
        font=cv2.FONT_HERSHEY_SIMPLEX,
        font_scale=font_scale * scale_factor,
        font_thickness=font_thickness,
        bg_color=bg_color,
        font_color=font_color,
        swatch_text_padding=int(dpg.get_value("swatch_text_padding") * scale_factor)
    )

    # Step 4: downsample with anti-aliasing
    preview_img = cv2.resize(
        hires_img,
        (hires_img.shape[1] // scale_factor, hires_img.shape[0] // scale_factor),
        interpolation=cv2.INTER_AREA
    )

    return preview_img

def update_output_preview(folder_path=None):
    try:
        if folder_path is None or not os.path.isdir(folder_path):
            img = np.full((500, 500, 3), 235, dtype=np.uint8)

            lines = [
                ("Preview unavailable", 0.9, (100, 100, 100), 2),
                ("Select a directory", 0.8, (120, 120, 120), 2),
                ("to generate a preview", 0.8, (120, 120, 120), 2),
            ]

            line_spacing = 50
            total_height = line_spacing * (len(lines) - 1)
            start_y = (img.shape[0] // 2) - (total_height // 2)

            for i, (text, scale, colour, thickness) in enumerate(lines):
                (text_w, text_h), _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness)
                x = (img.shape[1] - text_w) // 2
                y = start_y + i * line_spacing
                cv2.putText(img, text, (x, y), cv2.FONT_HERSHEY_SIMPLEX, scale, colour, thickness, cv2.LINE_AA)
        else:
            files = sorted([f for f in os.listdir(folder_path) if f.lower().endswith((".png",".jpg",".jpeg",".bmp",".tif",".tiff"))])

            if not files:
                return

            img_path = os.path.join(folder_path, files[0])

            img_color = cv2.imread(img_path)
            if img_color is None:
                return

            img_gray = cv2.cvtColor(img_color, cv2.COLOR_BGR2GRAY)

            if img_gray.shape[0] < CROP_SIZE or img_gray.shape[1] < CROP_SIZE:
                return

            crop = crop_image_by_position(img_gray, CROP_SIZE, get_value("crop_position_selector"))

            metrics = analyze_image(crop, GRID_SIZE)

            thick_bgr = (get_value("color_b_input_1"), get_value("color_g_input_1"), get_value("color_r_input_1"))
            even_bgr = (get_value("color_b_input_2"), get_value("color_g_input_2"), get_value("color_r_input_2"))
            thin_bgr = (get_value("color_b_input_3"), get_value("color_g_input_3"), get_value("color_r_input_3"))

            preview = create_colored_image(metrics["grid_means"], crop.shape, GRID_SIZE, metrics["mean"], thin_bgr, even_bgr, thick_bgr)

            preview = add_legend_below(
                preview,
                entries=[("Thin", thin_bgr), ("Even", even_bgr), ("Thick", thick_bgr)],
                pad=int(get_value("entry_padding")),
                swatch_size=int(get_value("swatch_size")),
                line_spacing=10,
                font=cv2.FONT_HERSHEY_SIMPLEX,
                font_scale=get_value("legend_font_size") / 32.0,
                font_thickness=get_value("legend_font_thickness"),
                bg_color=(get_value("legend_bg_b"), get_value("legend_bg_g"), get_value("legend_bg_r")),
                font_color=(get_value("legend_font_b"), get_value("legend_font_g"), get_value("legend_font_r")),
                swatch_text_padding=int(get_value("swatch_text_padding"))
            )

            img = cv2.resize(preview, (500, 500), interpolation=cv2.INTER_AREA)

        cv2.imwrite("preview_temp.png", img)

        width, height, channels, data = dpg.load_image("preview_temp.png")

        if dpg.does_item_exist("output_preview_image"):
            dpg.delete_item("output_preview_image")
        if dpg.does_item_exist("output_preview_texture"):
            dpg.delete_item("output_preview_texture")

        with dpg.texture_registry():
            dpg.add_static_texture(width, height, data, tag="output_preview_texture")

        dpg.add_image("output_preview_texture", parent="output_preview_group", tag="output_preview_image")

    except Exception as e:
        log(f"[Preview] {e}")

def clamp_and_update(tag, min_val, max_val):
    value = get_value(tag)
    clamped = max(min_val, min(max_val, value))
    if value != clamped:
        set_value(tag, clamped)
    update_font_preview()

def setup_gui():
    create_context()

    with window(label="Optical Uniformity", width=800, height=1050, no_collapse=True, no_close=True, no_move=True, no_resize=True, pos=(0, 0)):
        add_text("", tag="active_rgb_inputs", show=False)

        # Hidden RGB values for legend background color
        add_input_int(tag="legend_bg_r", width=0, default_value=255, show=False)
        add_input_int(tag="legend_bg_g", width=0, default_value=255, show=False)
        add_input_int(tag="legend_bg_b", width=0, default_value=255, show=False)

        # Hidden RGB values for legend font color
        add_input_int(tag="legend_font_r", width=0, default_value=0, show=False)
        add_input_int(tag="legend_font_g", width=0, default_value=0, show=False)
        add_input_int(tag="legend_font_b", width=0, default_value=0, show=False)

        with tab_bar(tag="main_tab_bar"):
            with tab(label="Settings"):
                add_text("Image Output Settings:")

                with table(header_row=True, width=770, borders_innerH=True, borders_outerH=True, borders_innerV=True, borders_outerV=True, policy=mvTable_SizingStretchProp):
                    add_table_column(label="Label");
                    add_table_column(label="R");
                    add_table_column(label="G");
                    add_table_column(label="B");
                    add_table_column(label="Preview");
                    add_table_column(label="Colour Picker")

                    with table_row():
                        add_text("Thick")
                        add_input_int(tag="color_r_input_1", width=60, default_value=255, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_1", "color_g_input_1", "color_b_input_1", "preview_1"))
                        add_input_int(tag="color_g_input_1", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_1", "color_g_input_1", "color_b_input_1", "preview_1"))
                        add_input_int(tag="color_b_input_1", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_1", "color_g_input_1", "color_b_input_1", "preview_1"))
                        add_button(label="", tag="preview_1", width=40, height=20)
                        add_button(label="Open Colour Palette", callback=open_color_picker, user_data=["color_r_input_1", "color_g_input_1", "color_b_input_1"])

                    with table_row():
                        add_text("Even")
                        add_input_int(tag="color_r_input_2", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_2", "color_g_input_2", "color_b_input_2", "preview_2"))
                        add_input_int(tag="color_g_input_2", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_2", "color_g_input_2", "color_b_input_2", "preview_2"))
                        add_input_int(tag="color_b_input_2", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_2", "color_g_input_2", "color_b_input_2", "preview_2"))
                        add_button(label="", tag="preview_2", width=40, height=20)
                        add_button(label="Open Colour Palette", callback=open_color_picker, user_data=["color_r_input_2", "color_g_input_2", "color_b_input_2"])

                    with table_row():
                        add_text("Thin")
                        add_input_int(tag="color_r_input_3", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_3", "color_g_input_3", "color_b_input_3", "preview_3"))
                        add_input_int(tag="color_g_input_3", width=60, default_value=0, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_3", "color_g_input_3", "color_b_input_3", "preview_3"))
                        add_input_int(tag="color_b_input_3", width=60, default_value=255, min_value=0, max_value=255, step=0, step_fast=0, callback=update_color_preview, user_data=("color_r_input_3", "color_g_input_3", "color_b_input_3", "preview_3"))
                        add_button(label="", tag="preview_3", width=40, height=20)
                        add_button(label="Open Colour Palette", callback=open_color_picker, user_data=["color_r_input_3", "color_g_input_3", "color_b_input_3"])

                add_separator()

                add_text("Output Settings:")
                add_checkbox(label="Create raw cropped images", tag="create_raw_checkbox", default_value=True)
                add_checkbox(label="Create contour images", tag="create_contour_checkbox", default_value=True)

                add_separator()

                with group(horizontal=True):
                    add_input_float(tag="swatch_size", width=100, min_value=1.0, format="%.1f", step=0.5, default_value=50.0, callback=lambda s, a, u: clamp_and_update("swatch_size", 0.0, 100.0))
                    add_text("Swatch Size")

                with group(horizontal=True):
                    add_input_float(tag="entry_padding", width=100, min_value=1.0, format="%.1f", step=0.5, default_value=15.0, callback=lambda s, a, u: clamp_and_update("entry_padding", 5.0, 100.0))
                    add_text("Vertical Padding")

                with group(horizontal=True):
                    add_input_float(tag="swatch_text_padding", width=100, min_value=1.0, format="%.1f", step=0.5, default_value=50.0, callback=lambda s, a, u: clamp_and_update("swatch_text_padding", 0.0, 100.0))
                    add_text("Horizontal Padding")

                with group(horizontal=True):
                    add_input_float(tag="legend_font_size", width=100, min_value=1.0, format="%.1f", step=0.1, default_value=42.0, callback=lambda s, a, u: clamp_and_update("legend_font_size", 10.0, 100.0))
                    add_text("Font Size")

                with group(horizontal=True):
                    add_input_int(tag="legend_font_thickness", width=100, min_value=1, step=1, default_value=2, callback=lambda s, a, u: clamp_and_update("legend_font_thickness", 0, 20))
                    add_text("Font Thickness")

                with group(horizontal=True):
                    add_button(label="", tag="legend_bg_preview", width=40, height=20)
                    add_button(label="Open Colour Palette", callback=open_color_picker, user_data=["legend_bg_r", "legend_bg_g", "legend_bg_b"])
                    add_text("Legend Background Colour")

                with group(horizontal=True):
                    add_button(label="", tag="legend_font_preview", width=40, height=20)
                    add_button(label="Open Colour Palette", callback=open_color_picker, user_data=["legend_font_r", "legend_font_g", "legend_font_b"])
                    add_text("Legend Font Colour")

                with group(horizontal=True):
                    add_text("Crop Position")
                    add_combo(tag="crop_position_selector", items=["Top-left", "Top-right", "Center", "Bottom-left", "Bottom-right"], default_value="Top-left", width=200, callback=lambda s,a,u: update_output_preview(last_selected_path))

                add_separator()

                add_text("Output Preview:")
                add_group(tag="output_preview_group")

                add_separator()

                with group():
                    add_button(label="Save Settings", callback=save_settings)
                    add_button(label="Analyse Images", tag="analyse_button_settings", callback=analyse_images_entry)

            with tab(label="Directories"):
                add_text("Choose directories for image analysis:")

                with table(tag="directory_table", header_row=True, width=770, height=248, scrollX=True, borders_innerH=True, borders_outerH=True, borders_innerV=True, borders_outerV=True, policy=mvTable_SizingStretchProp):
                    add_table_column(label="Folder")
                    add_table_column(label="Path")
                    add_table_column(label="Set")

                    for i in range(10):
                        folder_tag = f"folder_label_{i}"
                        path_tag = f"path_label_{i}"
                        with table_row():
                            add_input_text(tag=folder_tag, readonly=True, width=150)
                            add_input_text(tag=path_tag, readonly=True, width=500)
                            add_button(label="Set path", callback=open_folder_dialog, user_data={"path_tag": path_tag, "folder_tag": folder_tag})

                add_spacer()
                add_separator()
                add_spacer()

                add_button(label="Analyse Images", tag="analyse_button_directories", callback=analyse_images_entry)

                # Hidden file dialog
                add_file_dialog(directory_selector=True, show=False, callback=folder_selected, tag="folder_picker")

            with tab(label="Logging", tag="logging_tab"):
                add_child_window(tag="log_window", width=770, height=400, border=True)
                with group(horizontal=True):
                    add_button(label="Clear Logs", callback=lambda s, a, u: delete_item("log_window", children_only=True))
                    add_button(label="Save Log", callback=lambda s, a, u: save_log_to_file())

    create_viewport(title='Optical Uniformity', width=800, height=1050)
    setup_dearpygui()
    load_settings()

    # Set initial color previews for legend buttons
    update_color_preview(None, None, ("legend_bg_r", "legend_bg_g", "legend_bg_b", "legend_bg_preview"))
    update_color_preview(None, None, ("legend_font_r", "legend_font_g", "legend_font_b", "legend_font_preview"))

    # Draw the preview text
    def initialise_previews():
        update_font_preview()
        update_output_preview()

    dpg.set_frame_callback(1, initialise_previews)

    show_viewport()
    start_dearpygui()
    destroy_context()

def log(msg):
    if does_item_exist("log_window"):
        add_text(msg, parent="log_window", wrap=750)
    print(msg)

###### ANALYSIS HANDLING ######

# === Colorization ===
def create_colored_image(grid_means, shape, grid_size, mean_intensity, thin_bgr, even_bgr, thick_bgr):
    h, w = shape
    output_img = np.zeros((h, w, 3), dtype=np.uint8)

    lower = mean_intensity * 0.9
    upper = mean_intensity * 1.1

    i = 0
    for y in range(0, h, grid_size):
        for x in range(0, w, grid_size):
            if y + grid_size > h or x + grid_size > w:
                continue

            intensity = grid_means[i]
            i += 1

            if intensity < lower:
                color = thin_bgr
            elif intensity > upper:
                color = thick_bgr
            else:
                color = even_bgr

            output_img[y:y+grid_size, x:x+grid_size] = color

    return output_img

def add_legend_below(img, entries, pad=24, swatch_size=24, line_spacing=14, font=cv2.FONT_HERSHEY_SIMPLEX, font_scale=1.5, font_thickness=3, bg_color=(255, 255, 255), font_color=(0, 0, 0), swatch_text_padding=8):  # 👈 new param


    h, w = img.shape[:2]

    ascent = int(cv2.getTextSize("Ag", font, font_scale, font_thickness)[0][1])
    row_h = max(swatch_size, ascent) + line_spacing
    legend_h = pad + row_h + pad

    legend = np.full((legend_h, w, 3), bg_color, dtype=np.uint8)

    # layout
    blocks = []
    total_w = pad
    for label, color in entries:
        text_size, _ = cv2.getTextSize(label, font, font_scale, font_thickness)
        block_w = swatch_size + swatch_text_padding + text_size[0] + pad
        blocks.append((label, color, block_w, text_size))
        total_w += block_w
    total_w -= 24

    x = max(pad, (w - total_w) // 2)
    y_top = pad + (row_h - swatch_size) // 2
    text_baseline = pad + (row_h + ascent) // 2

    for label, color, block_w, text_size in blocks:
        cv2.rectangle(
            legend,
            (x, y_top),
            (x + swatch_size, y_top + swatch_size),
            color,
            thickness=-1,
            lineType=cv2.LINE_AA
        )
        cv2.putText(legend, label, (x + swatch_size + 8, text_baseline), font, font_scale, font_color, font_thickness, cv2.LINE_AA)
        x += block_w

    with_legend = np.vstack([img, legend])
    return with_legend

# === Central Crop ===
def crop_image(img, size):
    h, w = img.shape
    if h < size or w < size:
        raise ValueError("Image is smaller than crop size.")
    return img[:size, :size]

# === Analysis ===

# === Analysis Function ===
def analyze_image(img, grid_size):
    h, w = img.shape
    grid_means = []

    for y in range(0, h, grid_size):
        for x in range(0, w, grid_size):
            block = img[y:y+grid_size, x:x+grid_size]
            if block.shape[0] != grid_size or block.shape[1] != grid_size:
                continue
            grid_means.append(np.mean(block))

    grid_means = np.array(grid_means)
    mean_intensity = np.mean(grid_means)
    std_dev = np.std(grid_means)
    civ = (std_dev / mean_intensity) * 100 if mean_intensity != 0 else 0

    lower = mean_intensity * 0.9
    upper = mean_intensity * 1.1

    even = np.sum((grid_means >= lower) & (grid_means <= upper))
    thin = np.sum(grid_means < lower)
    thick = np.sum(grid_means > upper)
    total = len(grid_means)

    nec = (even / total) * 100
    thin_percent = (thin / total) * 100
    thick_percent = (thick / total) * 100

    return {
        "mean": mean_intensity,
        "std_dev": std_dev,
        "civ": civ,
        "thick%": thick_percent,
        "thin%": thin_percent,
        "nec": nec,
        "lower_thresh": lower,
        "upper_thresh": upper,
        "grid_means": grid_means,
        "shape": (h, w)
    }
def crop_image_by_position(img, size, position):
    h, w = img.shape[:2]
    if h < size or w < size:
        raise ValueError("Image is smaller than crop size.")

    if position == "Top-left":
        return img[0:size, 0:size]
    elif position == "Top-right":
        return img[0:size, w-size:w]
    elif position == "Bottom-left":
        return img[h-size:h, 0:size]
    elif position == "Bottom-right":
        return img[h-size:h, w-size:w]
    elif position == "Center":
        y_start = (h - size) // 2
        x_start = (w - size) // 2
        log(f"[DEBUG] Cropping {position}: y={y_start if 'y_start' in locals() else 'N/A'}, x={x_start if 'x_start' in locals() else 'N/A'}, size={size}")
        return img[y_start:y_start + size, x_start:x_start + size]
    else:
        # Fallback to top-left
        return img[0:size, 0:size]

def analyse_images():
    global analysis_running
    try:
        # Ensure resource directory exists
        os.makedirs(RESOURCE_DIR, exist_ok=True)

        # Gather folder paths from the GUI
        sample_folders = []
        for i in range(10):
            folder_tag = f"folder_label_{i}"
            path_tag = f"path_label_{i}"
            if does_item_exist(folder_tag) and does_item_exist(path_tag):
                folder_name = get_value(folder_tag)
                folder_path = get_value(path_tag)
                if folder_name and folder_path and os.path.isdir(folder_path):
                    dest_path = os.path.join(RESOURCE_DIR, folder_name)
                    os.makedirs(dest_path, exist_ok=True)

                    # Copy only image files to resources/folder_name
                    for f in os.listdir(folder_path):
                        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.bmp')):
                            src = os.path.join(folder_path, f)
                            dst = os.path.join(dest_path, f)
                            try:
                                if not os.path.exists(dst):
                                    shutil.copy2(src, dst)
                            except Exception as copy_err:
                                log(f"[WARNING] Failed to copy {f}: {copy_err}")

                    sample_folders.append(folder_name)

        if not sample_folders:
            log("[WARNING] No valid image folders were provided.")
            return

        # Check user settings for output options
        create_raw = get_value("create_raw_checkbox")
        create_contour = get_value("create_contour_checkbox")

        # read user-selected RGBs (Thick = row 1, Even = row 2, Thin = row 3)
        thick_rgb = (get_value("color_r_input_1"), get_value("color_g_input_1"), get_value("color_b_input_1"))
        even_rgb = (get_value("color_r_input_2"), get_value("color_g_input_2"), get_value("color_b_input_2"))
        thin_rgb = (get_value("color_r_input_3"), get_value("color_g_input_3"), get_value("color_b_input_3"))

        # OpenCV expects BGR
        thick_bgr = (thick_rgb[2], thick_rgb[1], thick_rgb[0])
        even_bgr = (even_rgb[2], even_rgb[1], even_rgb[0])
        thin_bgr = (thin_rgb[2], thin_rgb[1], thin_rgb[0])

        all_results = []

        for sample in sample_folders:
            log(f"\n[INFO] Processing sample folder: {sample}")
            sample_path = os.path.join(RESOURCE_DIR, sample)
            image_files = [f for f in os.listdir(sample_path)
                           if f.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.bmp'))]

            output_sample_dir = os.path.join(OUTPUT_DIR, sample)
            raw_sample_dir = os.path.join(OUTPUT_DIR, sample, "raw")
            os.makedirs(raw_sample_dir, exist_ok=True)
            os.makedirs(output_sample_dir, exist_ok=True)

            results = []

            for filename in image_files:
                img_path = os.path.join(sample_path, filename)

                img_color = cv2.imread(img_path)
                if img_color is None:
                    log(f"[WARNING] Could not read {filename}. Skipping.")
                    continue

                img_gray = cv2.cvtColor(img_color, cv2.COLOR_BGR2GRAY)

                if img_gray.shape[0] < CROP_SIZE or img_gray.shape[1] < CROP_SIZE:
                    log(f"[SKIP] {filename} too small for cropping.")
                    continue

                crop_position = get_value("crop_position_selector")
                log(f"[DEBUG] Crop position selected: {crop_position}")
                cropped = crop_image_by_position(img_gray, CROP_SIZE, crop_position)

                if create_raw:
                    raw_crop_color = crop_image_by_position(img_color, CROP_SIZE, crop_position)
                    raw_out_path = os.path.join(raw_sample_dir, f"{os.path.splitext(filename)[0]}_crop.png")
                    cv2.imwrite(raw_out_path, raw_crop_color)

                metrics = analyze_image(cropped, GRID_SIZE)
                if create_contour:
                    color_img = create_colored_image(metrics["grid_means"], cropped.shape, GRID_SIZE, metrics["mean"], thin_bgr, even_bgr, thick_bgr)

                    out_path = os.path.join(output_sample_dir, f"{os.path.splitext(filename)[0]}_{GRID_SIZE}x{GRID_SIZE}.png")
                    entries = [
                        ("Thin", thin_bgr),
                        ("Even", even_bgr),
                        ("Thick", thick_bgr),
                    ]
                    font_scale = get_value("legend_font_size") / 32.0
                    font_thickness = get_value("legend_font_thickness")
                    bg_color = (
                        get_value("legend_bg_b"),
                        get_value("legend_bg_g"),
                        get_value("legend_bg_r")
                    )
                    font_color = (
                        get_value("legend_font_b"),
                        get_value("legend_font_g"),
                        get_value("legend_font_r")
                    )
                    final_img = add_legend_below(
                        color_img,
                        entries=entries,
                        pad=int(get_value("entry_padding")),
                        swatch_size=int(get_value("swatch_size")),
                        line_spacing=10,
                        font=cv2.FONT_HERSHEY_SIMPLEX,
                        font_scale=font_scale,
                        font_thickness=font_thickness,
                        bg_color=bg_color,
                        font_color=font_color,
                        swatch_text_padding=int(get_value("swatch_text_padding"))
                    )

                    cv2.imwrite(out_path, final_img)

                results.append({
                    "Sample": filename,
                    "Mean Pixel Intensity": round(metrics["mean"], 2),
                    "CV Pixel Intensity (%)": round(metrics["civ"], 2),
                    "Thick Places (%)": round(metrics["thick%"], 2),
                    "Thick Threshold": round(metrics["upper_thresh"], 2),
                    "Thin Places (%)": round(metrics["thin%"], 2),
                    "Thin Threshold": round(metrics["lower_thresh"], 2),
                    "NEC (%)": round(metrics["nec"], 2),
                })

                log(f"[DONE] {filename}")

            if results:
                df = pd.DataFrame(results)

                # Append AVERAGE/STDEV formula rows
                avg_row = ["=AVERAGE(B2:B{})".format(len(df) + 1),
                           "=AVERAGE(C2:C{})".format(len(df) + 1),
                           "=AVERAGE(D2:D{})".format(len(df) + 1),
                           "=AVERAGE(E2:E{})".format(len(df) + 1),
                           "=AVERAGE(F2:F{})".format(len(df) + 1),
                           "=AVERAGE(G2:G{})".format(len(df) + 1),
                           "=AVERAGE(H2:H{})".format(len(df) + 1)]
                std_row = ["=STDEV(B2:B{})".format(len(df) + 1),
                           "=STDEV(C2:C{})".format(len(df) + 1),
                           "=STDEV(D2:D{})".format(len(df) + 1),
                           "=STDEV(E2:E{})".format(len(df) + 1),
                           "=STDEV(F2:F{})".format(len(df) + 1),
                           "=STDEV(G2:G{})".format(len(df) + 1),
                           "=STDEV(H2:H{})".format(len(df) + 1)]
                df.loc["Average"] = [""] + avg_row
                df.loc["StDev"] = [""] + std_row

                all_results.append((sample, df))
            else:
                log(f"[SKIP] No valid images found in folder: {sample}")

        if not all_results:
            log("[ERROR] No valid image data found in any folder.")
            return

        os.makedirs(EXCEL_PATH, exist_ok=True)
        timestamp = datetime.now().strftime("Output [%d-%m-%y] [%H.%M.%S].xlsx")
        excel_file_path = os.path.join(EXCEL_PATH, timestamp)

        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='w') as writer:
            for sample, df in all_results:
                df.to_excel(writer, sheet_name=sample, index=False, startrow=0)

                wb = writer.book
                ws = wb[sample]

                last_row = ws.max_row
                mean_row = last_row - 1
                sd_row = last_row

                bold_font = Font(bold=True)
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border

                ws.cell(row=mean_row, column=1).value = "Mean"
                ws.cell(row=mean_row, column=1).font = bold_font
                ws.cell(row=sd_row, column=1).value = "SD"
                ws.cell(row=sd_row, column=1).font = bold_font

                for col in range(1, ws.max_column + 1):
                    ws.cell(row=mean_row, column=col).font = bold_font
                    ws.cell(row=sd_row, column=col).font = bold_font

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column].width = adjusted_width

                log(f"[INFO] Sheet '{sample}' written.")

        log(f"\n All results saved to {excel_file_path}")

        # scroll to the bottom of the log window once analysis is fully complete
        try:
            if does_item_exist("log_window"):
                set_y_scroll("log_window", get_y_scroll_max("log_window"))
        except Exception:
            pass

        analysis_running = False
        enable_item("analyse_button_settings")
        enable_item("analyse_button_directories")

    except Exception as e:
        analysis_running = False
        enable_item("analyse_button_settings")
        enable_item("analyse_button_directories")
        log(f"[ERROR] Analysis failed: {e}")


###### SCRIPT RUN HANDLING ######

setup_gui()
